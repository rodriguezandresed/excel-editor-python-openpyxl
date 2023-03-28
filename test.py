from tkinter import E
from types import NoneType
from openpyxl import load_workbook
#Route to the file
workbook = load_workbook(filename="test.xlsx")
workbook.sheetnames


sheet = workbook.active

#  Here we specify the range for the excel file
for i in range (1,300):
    new_list1=[]
    new_list2=[]
    new_list3=[]
    new_list4=[]
    iterable=sheet.cell(row=i, column=1).value
    if type(iterable) is NoneType:
        mix=0
    else:
        mix=list(iterable)
        for x in range(0,len(mix)):
            if (0<=x<3):
                new_list1.append(mix[x])
                phrase1=''.join(new_list1)
                if phrase1 == "AL02":
                    final1 = "ALM-1002"
                else: 
                    final1 = "ALM-1003"
            elif (3<=x<6):
                new_list2.append(mix[x])
                phrase2=''.join(new_list2)               
            elif (6<=x<8):
                new_list3.append(mix[x])
                phrase3=''.join(new_list3)
            else: 
                if (mix[x] !="0" and x==10):
                    new_list4.append(mix[x])
                else:
                    new_list4.append(mix[x])
                phrase4=''.join(new_list4)  
        sheet.cell(row=i, column=2).value=final1+"-"+phrase2+"-"+phrase3+"-"+phrase4
workbook.save(filename="end_result.xlsx")
